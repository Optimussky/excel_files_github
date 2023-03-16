# ller_archivo_excel.py
# pip install openpyxl, tabulate

import openpyxl
from tabulate import tabulate

excel_dataframe = openpyxl.load_workbook("plantilla.xlsx")
dataframe = excel_dataframe.active

#print(dataframe)

data = []

for row in range(1, dataframe.max_row):
    #print(row)
    _row = [row,]
    for col in dataframe.iter_cols(1,dataframe.max_column-1):
        _row.append(col[row].value)

    data.append(_row)


headers = ['Id', 'Nombre','Puesto', 'Salario','Correo']
headers_align = (("center",) * 5)
#print(data)
print(tabulate(data, headers=headers, colalign=headers_align))
